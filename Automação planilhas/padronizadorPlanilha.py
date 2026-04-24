import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def novasColunas(df):
    df["CNPJ DO FORNECEDOR SÊNIOR"] = None
    df["CNPJ CLIENTE"] = None
    df["Cód. Cliente Sênior"] = None
    df["COD CENTRO DE CUSTO"] = None
    df["CONTA FINANCEIRA (REEMB. 1) (NÃO. 2)"] = None
    df["COD DO SERVIÇO (54321 - PF)     (12345 - PJ)"] = 940013

def reembolsavel(df):
    df.loc[df["Ato Reembolsável pelo Cliente?"] == "Não", "CONTA FINANCEIRA (REEMB. 1) (NÃO. 2)"] = 2
    df.loc[df["Ato Reembolsável pelo Cliente?"] == "Sim", "CONTA FINANCEIRA (REEMB. 1) (NÃO. 2)"] = 1 

def mapCC(df):

    mapeamentoCC = {
        "Centro de custo 1": 1,
        "Centro de custo 2": 2,
        "Centro de custo 3": 3
    }   

    df["COD CENTRO DE CUSTO"] = df["Centro de Custo"].map(mapeamentoCC)

def mapCorrespondente(df):

    mapeamentoCorrespondente = {
        "Correspondente 1": 111,
        "Correspondente 2": 222,
        "Correspondente 3": 333
    }

    df["CNPJ DO FORNECEDOR SÊNIOR"] = df["Correspondente"].map(mapeamentoCorrespondente)
    
def formatacaoDatas(planilha):
    wb = load_workbook(planilha)
    ws = wb.active
    
    colunasDatas = ["U", "V", "W", "X", "AG"]

    for datas in colunasDatas:
        for celula in ws[datas]:
            celula.number_format = 'dd/mm/yyyy'
    wb.save('PlanilhaFinal.xlsx')
    







    

